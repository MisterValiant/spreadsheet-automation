# Writing clean code

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# A resuable function is created so as to process multiple spreadsheets
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # For explanation purposes only
    # cell=sheet['a1']
    # cell=sheet.cell(1,1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * .9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # adding a chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')

# Read documentation to learn more